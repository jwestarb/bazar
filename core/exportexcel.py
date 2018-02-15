# -*- encoding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from django.http import HttpResponse


try:
    from django.utils import timezone
except ImportError:
    from datetime import datetime as timezone


class QuerysetToWorkbook(object):

    def __init__(self, queryset, columns, filename='report'):
        self.workbook = Workbook()
        self.sheet = self.workbook.get_active_sheet()
        self.sheet.title = "Dados Exportados"
        self.queryset = queryset
        self.columns = columns
        self.filename = filename
        self._row_num = 0

    def get_columns(self):
        return self.columns

    def render_column(self, row, column):
        """ Renders a column on a row
        """
        if hasattr(row, 'get_%s_display' % column):
            # It's a choice field
            text = getattr(row, 'get_%s_display' % column)()
        else:
            try:
                text = getattr(row, column)
            except AttributeError:
                obj = row
                for part in column.split('.'):
                    if obj is None:
                        break
                    obj = getattr(obj, part)

                text = obj

        return text

    def set_header(self):
        for col_num in range(len(self.columns)):
            c = self.sheet.cell(row=self._row_num + 1, column=col_num + 1)
            c.value = self.columns[col_num][0]
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color='A9D0F5', end_color='A9D0F5', fill_type='solid')
            self.sheet.column_dimensions[get_column_letter(col_num+1)].width = self.columns[col_num][1]

    def set_data(self):

        for obj in self.queryset:
            self._row_num += 1

            for col_num in range(len(self.columns)):
                c = self.sheet.cell(row=self._row_num + 1, column=col_num + 1)
                c.value = self.render_column(obj, self.columns[col_num][2])
                # c.style.alignment.wrap_text = True
        from django.db import connection
        print(len(connection.queries))
        for query in connection.queries:
            print('------------------------------------------------')
            print(query)

    def build_workbook(self):
        self.set_header()
        self.set_data()

    def response(self, filename=None):
        response = HttpResponse(content_type='application/octet-stream')
        self.workbook.save(response)
        response['Content-Disposition'] = 'attachment; filename=%s' % self.filename + '_' + \
                                          timezone.now().strftime('%d%m%Y_%H%M%S') + '.xlsx'
        return response
